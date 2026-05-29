import io
import csv
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    def _style_header(self, ws, row: int, cols: int):
        fill = PatternFill("solid", fgColor="1E40AF")
        font = Font(color="FFFFFF", bold=True)
        border = Border(
            bottom=Side(style="thin", color="000000"),
        )
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    def purchase_register(self, invoices: List) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Register"

        headers = [
            "Sr No", "Invoice Date", "Invoice Number", "Vendor Name", "Vendor GSTIN",
            "Customer GSTIN", "HSN/SAC", "Place of Supply", "State",
            "Taxable Amount", "CGST", "SGST", "IGST", "CESS", "Total Amount",
            "Status", "Client",
        ]
        ws.append(headers)
        self._style_header(ws, 1, len(headers))

        for i, inv in enumerate(invoices, 1):
            ws.append([
                i,
                str(inv.invoice_date or ""),
                str(inv.invoice_number or ""),
                str(inv.vendor_name or ""),
                str(inv.vendor_gstin or ""),
                str(inv.customer_gstin or ""),
                str(inv.hsn_sac or ""),
                str(inv.place_of_supply or ""),
                str(inv.state or ""),
                float(inv.taxable_amount or 0),
                float(inv.cgst or 0),
                float(inv.sgst or 0),
                float(inv.igst or 0),
                float(inv.cess or 0),
                float(inv.total_amount or 0),
                str(inv.status),
                inv.client.name if inv.client else "",
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def gst_upload_format(self, invoices: List) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "GSTR-2A Upload"

        headers = [
            "GSTIN of Supplier", "Invoice Number", "Invoice Date", "Invoice Value",
            "Place of Supply", "Reverse Charge", "Invoice Type",
            "Rate", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "CESS",
        ]
        ws.append(headers)
        self._style_header(ws, 1, len(headers))

        for inv in invoices:
            ws.append([
                str(inv.vendor_gstin or ""),
                str(inv.invoice_number or ""),
                str(inv.invoice_date or ""),
                float(inv.total_amount or 0),
                str(inv.place_of_supply or ""),
                "N",
                "Regular",
                "",
                float(inv.taxable_amount or 0),
                float(inv.igst or 0),
                float(inv.cgst or 0),
                float(inv.sgst or 0),
                float(inv.cess or 0),
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def csv_export(self, invoices: List) -> io.BytesIO:
        buf = io.BytesIO()
        wrapper = io.TextIOWrapper(buf, encoding="utf-8", newline="")
        writer = csv.writer(wrapper)

        writer.writerow([
            "ID", "Client", "Invoice Number", "Invoice Date", "Vendor Name", "Vendor GSTIN",
            "Customer GSTIN", "Taxable Amount", "CGST", "SGST", "IGST", "CESS", "Total Amount",
            "Status", "Uploaded At",
        ])

        for inv in invoices:
            writer.writerow([
                str(inv.id),
                inv.client.name if inv.client else "",
                inv.invoice_number or "",
                inv.invoice_date or "",
                inv.vendor_name or "",
                inv.vendor_gstin or "",
                inv.customer_gstin or "",
                str(inv.taxable_amount or ""),
                str(inv.cgst or ""),
                str(inv.sgst or ""),
                str(inv.igst or ""),
                str(inv.cess or ""),
                str(inv.total_amount or ""),
                inv.status,
                str(inv.created_at),
            ])

        wrapper.flush()
        buf.seek(0)
        return buf
